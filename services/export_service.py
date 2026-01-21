from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import GradeResult
from services.grading_service import GradingSession


class ExportService:
    """Service for exporting grading results."""

    def __init__(self):
        pass

    def export_to_excel(self, session: GradingSession, filepath: str,
                        include_feedback: bool = True) -> str:
        """
        Export grading results to Excel file.

        Args:
            session: The grading session with results
            filepath: Output file path
            include_feedback: Whether to include detailed feedback

        Returns:
            Path to the created file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"

        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(wrap_text=True, vertical='top')

        # Determine columns based on rubric elements
        if session.results and session.results[0].element_grades:
            element_names = [eg.element_name for eg in session.results[0].element_grades]
        elif session.rubric and session.rubric.elements:
            element_names = [e.name for e in session.rubric.elements]
        else:
            element_names = []

        # Build headers
        headers = ["#", "Student ID", "Student Name"]
        for name in element_names:
            headers.append(name)
        headers.extend(["Total", "Max Marks", "Percentage"])
        if include_feedback:
            headers.append("Feedback")
            if element_names:
                headers.append("Detailed Feedback")

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Write data
        for row_num, result in enumerate(session.results, 2):
            col = 1

            # Row number
            ws.cell(row=row_num, column=col, value=row_num - 1).border = thin_border
            col += 1

            # Student ID
            ws.cell(row=row_num, column=col, value=result.student_id).border = thin_border
            col += 1

            # Student Name
            ws.cell(row=row_num, column=col, value=result.student_name).border = thin_border
            col += 1

            # Element grades
            element_grades_dict = {eg.element_name: eg for eg in result.element_grades}
            for name in element_names:
                eg = element_grades_dict.get(name)
                marks = eg.marks_awarded if eg else 0
                cell = ws.cell(row=row_num, column=col, value=marks)
                cell.border = thin_border
                cell.alignment = center_align
                col += 1

            # Total
            cell = ws.cell(row=row_num, column=col, value=result.total_marks)
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = Font(bold=True)
            col += 1

            # Max Marks
            ws.cell(row=row_num, column=col, value=result.max_total_marks).border = thin_border
            col += 1

            # Percentage
            cell = ws.cell(row=row_num, column=col, value=f"{result.percentage:.1f}%")
            cell.border = thin_border
            cell.alignment = center_align

            # Color code based on percentage
            if result.percentage >= 70:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result.percentage >= 50:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            col += 1

            if include_feedback:
                # Overall feedback
                cell = ws.cell(row=row_num, column=col, value=result.overall_feedback)
                cell.border = thin_border
                cell.alignment = wrap_align
                col += 1

                # Detailed feedback per element
                if element_names:
                    detailed_feedback = []
                    for eg in result.element_grades:
                        if eg.feedback:
                            detailed_feedback.append(f"[{eg.element_name}]: {eg.feedback}")
                    cell = ws.cell(row=row_num, column=col, value="\n\n".join(detailed_feedback))
                    cell.border = thin_border
                    cell.alignment = wrap_align

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            if col <= 3:  # ID columns
                ws.column_dimensions[get_column_letter(col)].width = 15
            elif col <= len(element_names) + 3:  # Element columns
                ws.column_dimensions[get_column_letter(col)].width = 12
            elif col == len(headers) - 1 and include_feedback:  # Feedback column
                ws.column_dimensions[get_column_letter(col)].width = 40
            elif col == len(headers) and include_feedback:  # Detailed feedback
                ws.column_dimensions[get_column_letter(col)].width = 60
            else:
                ws.column_dimensions[get_column_letter(col)].width = 12

        # Add summary sheet
        self._add_summary_sheet(wb, session)

        # Save workbook
        wb.save(filepath)
        return filepath

    def _add_summary_sheet(self, wb: Workbook, session: GradingSession):
        """Add a summary statistics sheet."""
        ws = wb.create_sheet("Summary")

        # Styles
        header_font = Font(bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Calculate statistics
        if not session.results:
            ws.cell(row=1, column=1, value="No results available")
            return

        total_students = len(session.results)
        percentages = [r.percentage for r in session.results]
        avg_percentage = sum(percentages) / len(percentages) if percentages else 0
        min_percentage = min(percentages) if percentages else 0
        max_percentage = max(percentages) if percentages else 0

        # Grade distribution
        grade_dist = {
            'A (>=80%)': len([p for p in percentages if p >= 80]),
            'B (70-79%)': len([p for p in percentages if 70 <= p < 80]),
            'C (60-69%)': len([p for p in percentages if 60 <= p < 70]),
            'D (50-59%)': len([p for p in percentages if 50 <= p < 60]),
            'F (<50%)': len([p for p in percentages if p < 50])
        }

        # Write summary
        summary_data = [
            ("Grading Summary", ""),
            ("", ""),
            ("Total Students", total_students),
            ("Average Score", f"{avg_percentage:.1f}%"),
            ("Highest Score", f"{max_percentage:.1f}%"),
            ("Lowest Score", f"{min_percentage:.1f}%"),
            ("", ""),
            ("Grade Distribution", ""),
        ]

        for label, count in grade_dist.items():
            summary_data.append((label, count))

        summary_data.extend([
            ("", ""),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Rubric", session.rubric.name if session.rubric else "N/A"),
        ])

        for row_num, (label, value) in enumerate(summary_data, 1):
            cell1 = ws.cell(row=row_num, column=1, value=label)
            cell2 = ws.cell(row=row_num, column=2, value=value)

            if label in ["Grading Summary", "Grade Distribution"]:
                cell1.font = header_font

            cell1.border = thin_border
            cell2.border = thin_border

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def export_to_csv(self, session: GradingSession, filepath: str) -> str:
        """Export grading results to CSV file."""
        import csv

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Determine element names
            if session.results and session.results[0].element_grades:
                element_names = [eg.element_name for eg in session.results[0].element_grades]
            else:
                element_names = []

            # Write headers
            headers = ["Student ID", "Student Name"] + element_names + \
                      ["Total", "Max Marks", "Percentage", "Feedback"]
            writer.writerow(headers)

            # Write data
            for result in session.results:
                row = [result.student_id, result.student_name]

                element_grades_dict = {eg.element_name: eg for eg in result.element_grades}
                for name in element_names:
                    eg = element_grades_dict.get(name)
                    row.append(eg.marks_awarded if eg else 0)

                row.extend([
                    result.total_marks,
                    result.max_total_marks,
                    f"{result.percentage:.1f}%",
                    result.overall_feedback
                ])
                writer.writerow(row)

        return filepath

    def export_to_json(self, session: GradingSession, filepath: str) -> str:
        """Export grading results to JSON file."""
        data = {
            'session_id': session.id,
            'exported_at': datetime.now().isoformat(),
            'rubric': session.rubric.to_dict() if session.rubric else None,
            'results': [r.to_dict() for r in session.results],
            'summary': {
                'total_students': len(session.results),
                'average_percentage': sum(r.percentage for r in session.results) / len(session.results) if session.results else 0
            }
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        return filepath
